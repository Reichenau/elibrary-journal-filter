import os
import tkinter as tk
from tkinter import messagebox
import openpyxl
import subprocess


class JournalFilterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Фильтр журналов")
        
        # Соответствие между названиями в интерфейсе и значениями в таблице
        self.vak_mapping = {
            "к1": "1",
            "к2": "2", 
            "к3": "3",
            "без к": "без категории"
        }
        
        self.level_mapping = {
            "у1": "1",
            "у2": "2",
            "у3": "3",
            "у4": "4"
        }
        
        # Переменные для хранения состояний checkbox
        self.vak_vars = {k: tk.BooleanVar() for k in self.vak_mapping}
        self.level_vars = {k: tk.BooleanVar() for k in self.level_mapping}
        
        # Создание интерфейса
        self.create_widgets()
        
    def create_widgets(self):
        # Фрейм для категорий ВАК
        vak_frame = tk.LabelFrame(self.root, text="Категория ВАК", padx=5, pady=5)
        vak_frame.pack(padx=10, pady=5, fill=tk.X)
        
        for i, (k, var) in enumerate(self.vak_vars.items()):
            cb = tk.Checkbutton(vak_frame, text=k, variable=var)
            cb.grid(row=0, column=i, padx=5, pady=2, sticky=tk.W)
        
        # Фрейм для уровней белого списка
        level_frame = tk.LabelFrame(self.root, text="Уровень белого списка", padx=5, pady=5)
        level_frame.pack(padx=10, pady=5, fill=tk.X)
        
        for i, (lvl, var) in enumerate(self.level_vars.items()):
            cb = tk.Checkbutton(level_frame, text=lvl, variable=var)
            cb.grid(row=0, column=i, padx=5, pady=2, sticky=tk.W)
        
        # Кнопки (расположены как в предыдущей версии)
        button_frame = tk.Frame(self.root)
        button_frame.pack(padx=10, pady=10, fill=tk.X)
        
        get_btn = tk.Button(button_frame, text="Получить журналы", command=self.filter_journals)
        get_btn.pack(side=tk.LEFT, padx=5)
        
        update_btn = tk.Button(button_frame, text="Обновить журналы", command=self.update_journals)
        update_btn.pack(side=tk.LEFT, padx=5)
    
    def filter_journals(self):
        # Проверка выбора категорий
        if not any(var.get() for var in self.vak_vars.values()):
            messagebox.showwarning("Предупреждение", "Выберите хотя бы одну категорию ВАК")
            return
        
        if not any(var.get() for var in self.level_vars.values()):
            messagebox.showwarning("Предупреждение", "Выберите хотя бы один уровень белого списка")
            return
        
        # Проверка наличия файла
        if not os.path.exists("journals.xlsx"):
            messagebox.showwarning("Предупреждение", "Файл journals.xlsx не найден. Нажмите 'Обновить журналы'")
            return
        
        try:
            # Загрузка исходного файла
            wb = openpyxl.load_workbook("journals.xlsx")
            sheet = wb.active
            
            # Получаем выбранные значения
            selected_vak = [self.vak_mapping[k] for k, var in self.vak_vars.items() if var.get()]
            selected_levels = [self.level_mapping[k] for k, var in self.level_vars.items() if var.get()]
            
            # Создаем новый файл для результатов
            result_wb = openpyxl.Workbook()
            result_sheet = result_wb.active
            result_sheet.append(["Название журнала", "Ссылка на журнал", "Категория ВАК", "Уровень белого списка"])
            
            # Обрабатываем каждую строку
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) < 4:  # Проверяем, что в строке достаточно данных
                    continue
                
                name, link, vak, level = row
                
                # Обработка значений None
                vak = str(vak) if vak is not None else "без категории"
                level = str(level) if level is not None else ""
                
                # Проверка соответствия фильтрам
                vak_ok = ("без категории" in selected_vak and vak == "без категории") or (vak in selected_vak)
                level_ok = level in selected_levels
                
                if vak_ok and level_ok:
                    result_sheet.append([name, link, vak, level])
            
            # Сохраняем результат
            result_wb.save("filtered_journals.xlsx")
            found_count = result_sheet.max_row - 1
            if found_count > 0:
                messagebox.showinfo("Успех", f"Найдено {found_count} журналов. Результат сохранен в filtered_journals.xlsx")
            else:
                messagebox.showwarning("Результат", "Журналы по выбранным критериям не найдены")
            
        except Exception as e:
            messagebox.showerror("Ошибка", f"Ошибка при обработке файла: {str(e)}")
    
    def update_journals(self):
        try:
            subprocess.run(["python", "update.py"], check=True)
            messagebox.showinfo("Успех", "Журналы успешно обновлены")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось обновить журналы: {str(e)}")


if __name__ == "__main__":
    root = tk.Tk()
    app = JournalFilterApp(root)
    root.mainloop()
